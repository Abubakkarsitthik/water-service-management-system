from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from app.database.mongodb import get_database
from app.core.security import get_current_user
from app.schemas.customer import CustomerCreate, CustomerUpdate
from bson import ObjectId
from datetime import datetime, timezone
import math
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List
import calendar

router = APIRouter(prefix="/customers", tags=["Customers"])


def format_customer(customer: dict) -> dict:
    return {
        "id": str(customer["_id"]),
        "customer_id": customer.get("customer_id", ""),
        "name": customer.get("name", ""),
        "phone": customer.get("phone", customer.get("mobile", "")),
        "alternate_phone": customer.get("alternate_phone", customer.get("alternate_mobile")),
        "address": customer.get("address"),
        "service_type": customer.get("service_type", ""),
        "installation_date": customer.get("installation_date", ""),
        "next_reminder_date": customer.get("next_reminder_date"),
        "notes": customer.get("notes"),
        "created_at": customer.get("created_at"),
        "updated_at": customer.get("updated_at"),
    }


def normalize_date(date_str: str) -> str:
    """Convert any common date format to YYYY-MM-DD. Returns empty string if unable."""
    if not date_str:
        return ""
    date_str = str(date_str).strip()
    # Already YYYY-MM-DD
    if len(date_str) >= 10 and date_str[4] == '-':
        return date_str[:10]
    # Try multiple formats
    formats = [
        "%d/%m/%Y",  # 13/01/2026 (Indian)
        "%m/%d/%Y",  # 01/13/2026 (US)
        "%d-%m-%Y",  # 13-01-2026
        "%m-%d-%Y",  # 01-13-2026
        "%d.%m.%Y",  # 13.01.2026
        "%Y/%m/%d",  # 2026/01/13
        "%d %b %Y",  # 13 Jan 2026
        "%d-%b-%Y",  # 13-Jan-2026
        "%d/%b/%Y",  # 13/Jan/2026
        "%b %d, %Y", # Jan 13, 2026
    ]
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Last resort: try dateutil if available
    try:
        from dateutil import parser as dateutil_parser
        dt = dateutil_parser.parse(date_str, dayfirst=True)
        return dt.strftime("%Y-%m-%d")
    except Exception:
        pass
    return ""


def add_months(date_str: str, months: int) -> str:
    """Add N months to a date string (any format), preserving the day."""
    if not date_str:
        return ""
    normalized = normalize_date(date_str)
    if not normalized:
        return ""
    try:
        dt = datetime.strptime(normalized, "%Y-%m-%d")
        new_month = dt.month + months
        year_offset = (new_month - 1) // 12
        new_year = dt.year + year_offset
        new_month = (new_month - 1) % 12 + 1
        max_day = calendar.monthrange(new_year, new_month)[1]
        new_day = min(dt.day, max_day)
        return datetime(new_year, new_month, new_day).strftime("%Y-%m-%d")
    except Exception:
        return ""


async def get_interval_months(db, service_type_name: str) -> int:
    """Look up the reminder interval (in months) for a service type. Returns None if not found."""
    if not service_type_name:
        return None
    st = await db.service_types.find_one(
        {"name": {"$regex": f"^{service_type_name}$", "$options": "i"}}
    )
    if not st:
        return None
    # Prefer reminder_interval_months; fall back to computing from default_interval_days
    months = st.get("reminder_interval_months")
    if months is None:
        days = st.get("default_interval_days")
        if days:
            months = round(days / 30)
    return months


@router.post("/create", response_model=dict)
async def create_customer(
    customer: CustomerCreate,
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    count = await db.customers.count_documents({})
    customer_id = f"CUS-{count + 1:05d}"

    # Always auto-calculate next_reminder_date from installation_date + service interval
    next_reminder = customer.next_reminder_date
    if not next_reminder:
        if customer.installation_date and customer.service_type:
            months = await get_interval_months(db, customer.service_type)
            if months:
                next_reminder = add_months(customer.installation_date, months)

    customer_doc = {
        "customer_id": customer_id,
        "name": customer.name,
        "phone": customer.phone,
        "alternate_phone": customer.alternate_phone,
        "address": customer.address,
        "service_type": customer.service_type,
        "installation_date": customer.installation_date,
        "next_reminder_date": next_reminder,
        "notes": customer.notes,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }

    result = await db.customers.insert_one(customer_doc)
    return {
        "message": "Customer created successfully",
        "customer_id": customer_id,
        "id": str(result.inserted_id),
        "next_reminder_date": next_reminder,
    }


@router.get("/list", response_model=dict)
async def list_customers(
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    search: str = Query(None),
    service_type: str = Query(None),
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    query = {}

    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"mobile": {"$regex": search, "$options": "i"}},
            {"customer_id": {"$regex": search, "$options": "i"}},
            {"address": {"$regex": search, "$options": "i"}},
        ]

    if service_type:
        query["service_type"] = {"$regex": f"^{service_type}$", "$options": "i"}

    total = await db.customers.count_documents(query)
    skip = (page - 1) * limit

    customers_cursor = db.customers.find(query).sort("created_at", -1).skip(skip).limit(limit)
    customers = []
    async for customer in customers_cursor:
        customers.append(format_customer(customer))

    return {
        "data": customers,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": math.ceil(total / limit) if total > 0 else 0,
    }


@router.get("/details/{customer_id}", response_model=dict)
async def get_customer(
    customer_id: str,
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    customer = None
    if ObjectId.is_valid(customer_id):
        customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
    if not customer:
        customer = await db.customers.find_one({"customer_id": customer_id})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    result = format_customer(customer)

    # Attach reminder history
    reminders = []
    cursor = db.reminders.find({"customer_id": str(customer["_id"])}).sort("created_at", -1).limit(10)
    async for r in cursor:
        reminders.append({
            "id": str(r["_id"]),
            "message": r.get("message", ""),
            "channel": r.get("channel", "whatsapp"),
            "created_at": r.get("created_at", ""),
        })
    result["reminders"] = reminders
    return result


@router.put("/update/{customer_id}", response_model=dict)
async def update_customer(
    customer_id: str,
    update_data: CustomerUpdate,
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    if not ObjectId.is_valid(customer_id):
        raise HTTPException(status_code=400, detail="Invalid customer ID")

    # Get current customer to check what changed
    existing = await db.customers.find_one({"_id": ObjectId(customer_id)})
    if not existing:
        raise HTTPException(status_code=404, detail="Customer not found")

    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}

    # Auto-recalculate next_reminder_date if installation_date or service_type changed
    # and no explicit next_reminder_date was provided
    new_install = update_dict.get("installation_date") or existing.get("installation_date")
    new_service = update_dict.get("service_type") or existing.get("service_type")
    changed = (
        "installation_date" in update_dict or "service_type" in update_dict
    )
    if changed and "next_reminder_date" not in update_dict:
        months = await get_interval_months(db, new_service)
        if months and new_install:
            update_dict["next_reminder_date"] = add_months(new_install, months)

    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()

    result = await db.customers.update_one(
        {"_id": ObjectId(customer_id)}, {"$set": update_dict}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")

    return {
        "message": "Customer updated successfully",
        "next_reminder_date": update_dict.get("next_reminder_date"),
    }


@router.delete("/delete/{customer_id}", response_model=dict)
async def delete_customer(
    customer_id: str,
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    if not ObjectId.is_valid(customer_id):
        raise HTTPException(status_code=400, detail="Invalid customer ID")

    result = await db.customers.delete_one({"_id": ObjectId(customer_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")

    return {"message": "Customer deleted successfully"}


@router.post("/complete-service/{customer_id}", response_model=dict)
async def complete_service(
    customer_id: str,
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    if not ObjectId.is_valid(customer_id):
        raise HTTPException(status_code=400, detail="Invalid customer ID")

    customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    service_type_name = customer.get("service_type")
    months_to_add = await get_interval_months(db, service_type_name)
    if not months_to_add:
        months_to_add = 3  # Hard fallback

    # Advance strictly from the previous scheduled date (not today) to keep the schedule aligned
    base_date = (
        customer.get("next_reminder_date")
        or customer.get("installation_date")
        or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    )
    new_next_date = add_months(base_date, months_to_add)

    await db.customers.update_one(
        {"_id": ObjectId(customer_id)},
        {"$set": {
            "next_reminder_date": new_next_date,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
    )

    # Log reminder completion
    await db.reminders.insert_one({
        "customer_id": str(customer["_id"]),
        "customer_name": customer.get("name", ""),
        "message": f"Service completed. Next reminder scheduled for {new_next_date}.",
        "reminder_type": "completed",
        "channel": "manual",
        "status": "completed",
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    return {
        "message": "Service completed. Next reminder scheduled.",
        "next_reminder_date": new_next_date,
    }


@router.post("/import", response_model=dict)
async def import_customers(
    customers_data: List[dict],
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    """Import customers from parsed Excel/CSV data (array of objects)."""
    inserted = 0
    errors = []

    for i, row in enumerate(customers_data):
        try:
            name = str(row.get("name") or row.get("Name") or "").strip()
            phone = str(
                row.get("phone") or row.get("Phone") or
                row.get("mobile") or row.get("Mobile") or ""
            ).strip()

            if not name or not phone:
                errors.append(f"Row {i+1}: Name and Phone are required.")
                continue

            installation_date = str(
                row.get("installation_date") or row.get("Installation Date") or ""
            ).strip() or None
            service_type = str(
                row.get("service_type") or row.get("Service Type") or ""
            ).strip() or None
            next_reminder_raw = str(
                row.get("next_reminder_date") or row.get("Next Reminder Date") or ""
            ).strip() or None

            # Validate required fields
            if not installation_date:
                errors.append(f"Row {i+1} ({name}): Installation Date is required.")
                continue
            if not service_type:
                errors.append(f"Row {i+1} ({name}): Service Type is required.")
                continue

            # Normalize the installation date from any Excel format to YYYY-MM-DD
            installation_date = normalize_date(installation_date)
            if not installation_date:
                errors.append(f"Row {i+1} ({name}): Installation Date format not recognized. Use YYYY-MM-DD or DD/MM/YYYY.")
                continue

            # Normalize next_reminder if provided
            next_reminder = normalize_date(next_reminder_raw) if next_reminder_raw else None

            # Auto-calculate next_reminder if not provided
            if not next_reminder:
                months = await get_interval_months(db, service_type)
                if months is None:
                    errors.append(
                        f"Row {i+1} ({name}): Service Type '{service_type}' not found. "
                        f"Create it in Service Types first."
                    )
                    continue
                next_reminder = add_months(installation_date, months)

            count = await db.customers.count_documents({})
            customer_id = f"CUS-{count + inserted + 1:05d}"

            doc = {
                "customer_id": customer_id,
                "name": name,
                "phone": phone,
                "alternate_phone": str(row.get("alternate_phone") or row.get("Alternate Phone") or "").strip() or None,
                "address": str(row.get("address") or row.get("Address") or "").strip() or None,
                "service_type": service_type,
                "installation_date": installation_date,
                "next_reminder_date": next_reminder,
                "notes": str(row.get("notes") or row.get("Notes") or "").strip() or None,
                "created_by": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.customers.insert_one(doc)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)}")

    return {
        "message": f"Import complete. {inserted} imported, {len(errors)} failed.",
        "inserted": inserted,
        "errors": errors,
    }


@router.get("/export")
async def export_customers(
    service_type: str = Query(None),
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    """Export all customers as an Excel file."""
    query = {}
    if service_type:
        query["service_type"] = {"$regex": f"^{service_type}$", "$options": "i"}

    cursor = db.customers.find(query).sort("created_at", -1)
    customers = []
    async for c in cursor:
        customers.append(format_customer(c))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        "Customer ID", "Name", "Phone", "Alternate Phone",
        "Address", "Service Type", "Installation Date",
        "Next Reminder Date", "Notes", "Created At"
    ]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 18)

    for row_num, c in enumerate(customers, 2):
        ws.cell(row=row_num, column=1, value=c["customer_id"])
        ws.cell(row=row_num, column=2, value=c["name"])
        ws.cell(row=row_num, column=3, value=c["phone"])
        ws.cell(row=row_num, column=4, value=c.get("alternate_phone") or "")
        ws.cell(row=row_num, column=5, value=c.get("address") or "")
        ws.cell(row=row_num, column=6, value=c.get("service_type") or "")
        ws.cell(row=row_num, column=7, value=c.get("installation_date") or "")
        ws.cell(row=row_num, column=8, value=c.get("next_reminder_date") or "")
        ws.cell(row=row_num, column=9, value=c.get("notes") or "")
        ws.cell(row=row_num, column=10, value=c.get("created_at", "")[:10] if c.get("created_at") else "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
