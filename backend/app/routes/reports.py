from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from app.database.mongodb import get_database
from app.core.security import get_current_user
from datetime import datetime, timezone, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter(prefix="/reports", tags=["Reports"])


def make_excel(headers: list, rows: list[list], sheet_title: str = "Report") -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 15)

    for row_num, row in enumerate(rows, 2):
        for col_num, val in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=str(val) if val is not None else "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/customers")
async def customer_report(
    format: str = Query("json", regex="^(json|excel)$"),
    service_type: str = Query(None),
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    query = {}
    if service_type:
        query["service_type"] = service_type

    cursor = db.customers.find(query).sort("created_at", -1)
    data = []
    async for c in cursor:
        data.append({
            "Customer ID": c.get("customer_id", ""),
            "Name": c.get("name", ""),
            "Phone": c.get("phone", c.get("mobile", "")),
            "Alternate Phone": c.get("alternate_phone", c.get("alternate_mobile", "")),
            "Address": c.get("address", ""),
            "Service Type": c.get("service_type", c.get("installation", {}).get("service_type", "") if c.get("installation") else ""),
            "Installation Date": c.get("installation_date", c.get("installation", {}).get("installation_date", "") if c.get("installation") else ""),
            "Next Reminder Date": c.get("next_reminder_date", ""),
            "Notes": c.get("notes", ""),
            "Created At": c.get("created_at", "")[:10] if c.get("created_at") else "",
        })

    if format == "excel":
        if not data:
            headers = ["Customer ID", "Name", "Phone", "Alternate Phone", "Address",
                       "Service Type", "Installation Date", "Next Reminder Date", "Notes", "Created At"]
            rows = []
        else:
            headers = list(data[0].keys())
            rows = [list(row.values()) for row in data]
        output = make_excel(headers, rows, "Customer Report")
        filename = f"customer_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {"data": data, "total": len(data)}


@router.get("/due-services")
async def due_service_report(
    format: str = Query("json", regex="^(json|excel)$"),
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    """Customers where next_reminder_date is today or in the past (overdue) or upcoming 30 days."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    month_end = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")

    query = {"next_reminder_date": {"$lte": month_end}}
    cursor = db.customers.find(query).sort("next_reminder_date", 1)
    data = []
    async for c in cursor:
        reminder_date = c.get("next_reminder_date", "")
        status = "Overdue" if reminder_date and reminder_date < today else "Upcoming"
        data.append({
            "Customer ID": c.get("customer_id", ""),
            "Name": c.get("name", ""),
            "Phone": c.get("phone", c.get("mobile", "")),
            "Service Type": c.get("service_type", c.get("installation", {}).get("service_type", "") if c.get("installation") else ""),
            "Next Reminder Date": reminder_date,
            "Status": status,
            "Address": c.get("address", ""),
        })

    if format == "excel":
        headers = list(data[0].keys()) if data else ["Customer ID", "Name", "Phone", "Service Type", "Next Reminder Date", "Status", "Address"]
        rows = [list(row.values()) for row in data]
        output = make_excel(headers, rows, "Due Services")
        filename = f"due_services_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {"data": data, "total": len(data)}


@router.get("/reminders")
async def reminder_report(
    format: str = Query("json", regex="^(json|excel)$"),
    db=Depends(get_database),
    current_user: dict = Depends(get_current_user),
):
    cursor = db.reminders.find().sort("created_at", -1).limit(500)
    data = []
    async for r in cursor:
        data.append({
            "Customer Name": r.get("customer_name", ""),
            "Channel": r.get("channel", "whatsapp"),
            "Status": r.get("status", ""),
            "Message Preview": (r.get("message", "")[:80] + "...") if len(r.get("message", "")) > 80 else r.get("message", ""),
            "Sent At": r.get("created_at", "")[:19] if r.get("created_at") else "",
        })

    if format == "excel":
        headers = list(data[0].keys()) if data else ["Customer Name", "Channel", "Status", "Message Preview", "Sent At"]
        rows = [list(row.values()) for row in data]
        output = make_excel(headers, rows, "Reminders")
        filename = f"reminder_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    return {"data": data, "total": len(data)}
